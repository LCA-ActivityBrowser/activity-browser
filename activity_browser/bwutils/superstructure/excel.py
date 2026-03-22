# -*- coding: utf-8 -*-
from ast import literal_eval
from pathlib import Path
from typing import List, Union
from logging import getLogger

import openpyxl
import pandas as pd

from .utils import SUPERSTRUCTURE

log = getLogger(__name__)


def convert_tuple_str(x):
    try:
        return literal_eval(x)
    except (ValueError, SyntaxError) as e:
        return x


def get_sheet_names(document_path: Union[str, Path]) -> List[str]:
    try:
        wb = openpyxl.load_workbook(filename=document_path, read_only=True)
        return wb.sheetnames
    except UnicodeDecodeError as e:
        log.error("Given document uses an unknown encoding: {}".format(e))


def get_header_index(document_path: Union[str, Path], import_sheet: int):
    """Retrieves the line index for the column headers, will raise an
    exception if not found in the first 10 rows.
    """
    try:
        wb = openpyxl.load_workbook(filename=document_path, read_only=True)
        sheet = wb.worksheets[import_sheet]
        for i in range(10):
            value = sheet.cell(i + 1, 1).value
            if isinstance(value, str):
                wb.close()
                return i
    except IndexError as e:
        wb.close()
        raise IndexError("Expected headers not found in file").with_traceback(
            e.__traceback__
        )
    except UnicodeDecodeError as e:
        log.error("Given document uses an unknown encoding: {}".format(e))
        wb.close()
    raise ValueError("Could not find required headers in given document sheet.")


def valid_cols(name: str) -> bool:
    """Callable which evaluates if a specific column should be used."""
    return False if str(name).startswith("#") else True


def import_from_excel(
    document_path: Union[str, Path], import_sheet: int = 1
) -> pd.DataFrame:
    """Import all of the exchanges and their scenario amounts from a given
    document and sheet index.

    The default index chosen represents the second sheet (first after the
    'information' sheet).

    Any '*' character used at the start of a row or will cause that row
    to be excluded from the import.
    A '#' character at the start of a column will cause that column to be
    excluded from the import.

    'usecols' is used to exclude specific columns from the excel document.
    'comment' is used to exclude specific rows from the excel document.
    """
    data = pd.DataFrame({})
    try:
        header_idx = get_header_index(document_path, import_sheet)
        data = pd.read_excel(
            document_path,
            sheet_name=import_sheet,
            header=header_idx,
            usecols=valid_cols,
            comment="*",
            na_values="",
            keep_default_na=False,
            engine="openpyxl",
        )
        diff = SUPERSTRUCTURE.difference(data.columns)
        if not diff.empty:
            raise ValueError(
                "Missing required column(s) for superstructure: {}".format(
                    diff.to_list()
                )
            )

        # Convert specific columns that may have tuples as strings
        columns = ["from categories", "from key", "to categories", "to key"]
        data.loc[:, columns] = data[columns].map(convert_tuple_str)
    except:
        # skip the error checks here, these now occur in the calling layout.tabs.LCA_setup module
        pass
    return data
