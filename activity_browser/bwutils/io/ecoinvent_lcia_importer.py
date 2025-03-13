import functools
import warnings
import tqdm
from numbers import Number

from bw2data import Database, config, methods, Method
from openpyxl import load_workbook

from bw2io.strategies import (
    drop_unspecified_subcategories,
    link_iterable_by_fields,
    normalize_units,
    rationalize_method_names,
    set_biosphere_type,
)
from bw2io.importers.base_lcia import LCIAImporter


class EcoinventLCIAImporter(LCIAImporter):
    """
    A class for importing ecoinvent-compatible LCIA methods

    """
    def __init__(self, filepath, biosphere=None):
        self.strategies = []
        self.applied_strategies = []
        self.filepath = filepath
        self.biosphere_name = biosphere

        if self.biosphere_name:
            self.set_biosphere(self.biosphere_name)

    @classmethod
    def setup_with_ei_excel(cls, file: str, biosphere_database: str | None = None):
        """Initialize an instance of EcoinventLCIAImporter.

        Defines strategies in ``__init__`` because ``config.biosphere`` is dynamic.
        """
        importer = cls(file, biosphere_database)
        importer.strategies = [
            normalize_units,
            set_biosphere_type,
            drop_unspecified_subcategories,
            functools.partial(
                link_iterable_by_fields,
                other=Database(biosphere_database or config.biosphere),
                fields=("name", "categories"),
            ),
        ]
        importer.cf_data, importer.units = convert_lcia_methods_data(file)
        importer.separate_methods()
        return importer

    def set_biosphere(self, biosphere_database: str):
        self.strategies = [
            normalize_units,
            set_biosphere_type,
            drop_unspecified_subcategories,
            functools.partial(
                link_iterable_by_fields,
                other=Database(biosphere_database),
                fields=("name", "categories"),
                relink=True,
            ),
        ]

    def add_rationalize_method_names_strategy(self):
        self.strategies.append(rationalize_method_names)

    def separate_methods(self):
        """Separate the list of CFs into distinct methods"""
        methods = {obj["method"] for obj in self.cf_data}

        self.data = {}

        missing = set()

        for line in self.cf_data:
            if line["method"] not in self.units:
                missing.add(line["method"])

        if missing:
            _ = lambda x: sorted([str(y) for y in x])
            warnings.warn("Missing units for following:" + " | ".join(_(missing)))

        for line in self.cf_data:
            assert isinstance(line["amount"], Number)

            if line["method"] not in self.data:
                self.data[line["method"]] = {
                    "filename": self.filepath,
                    "unit": self.units.get(line["method"], ""),
                    "name": line["method"],
                    "description": "",
                    "exchanges": [],
                }

            self.data[line["method"]]["exchanges"].append(
                {
                    "name": line["name"],
                    "categories": line["categories"],
                    "amount": line["amount"],
                }
            )

        self.data = list(self.data.values())

    def apply_strategies(self, strategies=None, verbose=False):
        strategies = strategies or self.strategies
        for strategy in tqdm.tqdm(strategies, desc="Applying strategies", total=len(strategies)):
            self.apply_strategy(strategy, verbose=verbose)

    def prepend_methods(self, prepend: str):
        if not prepend:
            return
        for method in tqdm.tqdm(self.data, desc=f"Prepending {prepend} to ICs"):
            method["name"] = tuple([prepend, *method["name"]])

    def write_methods(self, overwrite=False, verbose=True):
        num_methods, num_cfs, num_unlinked = self.statistics(False)
        if num_unlinked:
            raise ValueError(
                ("Can't write unlinked methods ({} unlinked cfs)").format(num_unlinked)
            )
        for ds in tqdm.tqdm(self.data, total=len(self.data), desc="Processing CF's"):
            if ds["name"] in methods:
                if overwrite:
                    del methods[ds["name"]]
                else:
                    raise ValueError(
                        (
                            "Method {} already exists. Use "
                            "``overwrite=True`` to overwrite existing methods"
                        ).format(ds["name"])
                    )
            method = Method(ds["name"])
            method.register(
                description=ds["description"],
                filename=ds["filename"],
                unit=ds["unit"],
            )
            method.write(self._reformat_cfs(ds["exchanges"]))
        if verbose:
            print(
                "Wrote {} LCIA methods with {} characterization factors".format(
                    num_methods, num_cfs
                )
            )


def convert_lcia_methods_data(filename: str):
    sheet = load_workbook(filename, read_only=True)["CFs"]

    def process_row(row):
        data = [cell.value for i, cell in zip(range(8), row)]
        if not isinstance(data[-1], Number):
            return None
        else:
            return {
                "method": tuple(data[:3]),
                "name": data[3],
                "categories": tuple(data[4:6]),
                "amount": data[6],
            }

    cf_data = []
    for rowidx, row in tqdm.tqdm(enumerate(sheet.rows), total=sheet.max_row, desc="Processing CF's"):
        if rowidx:
            cf_data.append(process_row(row))

    sheet = load_workbook(filename, read_only=True)["Indicators"]

    def process_unit_row(row):
        data = [cell.value for i, cell in zip(range(4), row)]
        return tuple(data[:3]), data[3]

    units = {}
    for rowidx, row in tqdm.tqdm(enumerate(sheet.rows), total=sheet.max_row, desc="Processing indicators"):
        if rowidx:
            key, value = process_unit_row(row)
            units[key] = value

    return cf_data, units
